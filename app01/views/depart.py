from django.shortcuts import render, redirect, HttpResponse
from app01 import models
from django.core.files.uploadedfile import InMemoryUploadedFile
from app01.utils.pagination import Pagination
from openpyxl import load_workbook


# Create your views here.
def depart_list(request):
    """部门列表"""

    # 去数据库中获取所有部门信息
    # 得到Query Set
    queryset = models.Department.objects.all()
    page_obj = Pagination(request, queryset)
    context = {
        "page_string": page_obj.html(),
        "queryset": page_obj.page_queryset,

    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加部门"""

    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取post请求的数据(不考虑title为空)

    title = request.POST.get("title")
    models.Department.objects.create(title=title)

    # 重定向
    return redirect('/depart/list/')


def depart_delete(request):
    """删除部门"""
    # 获取id
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()

    # 重定向
    return redirect('/depart/list/')


def depart_edit(request, nid):
    """修改部门"""
    if request.method == "GET":
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {'title': row_object.title})
    # 根据nid获取数据 并更新数据库内容
    models.Department.objects.filter(id=nid).update(title=request.POST.get('title'))

    # 重定向
    return redirect('/depart/list/')


def depart_multi(request):
    """批量上传（excel）"""

    # work_book_object = load_workbook('文件路径')
    # sheet = work_book_object.worksheets[0]
    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    # 对象传递给openpyxl，读取内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[3].value
        print(text)
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    # cell = sheet.cell(row=1, column=1)  # 读取第一行第一列
    # print(cell.value)
    return redirect('/depart/list/')
